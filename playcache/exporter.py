"""Export the SQLite catalog to an .xlsx file matching the 6-column reference layout."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .db import Database

HEADERS = [
    "GAME NAME", "PLATFORM", "GOG / STEAM", "USER RATING",
    "GAME TYPE", "SHORT DESCRIPTION",
]

_COL_WIDTHS = dict(zip(HEADERS, [42, 16, 16, 14, 24, 90], strict=True))

_INJECTION_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _sanitize_cell(val):
    """Prevent CSV/Excel formula injection.

    If a string value starts with a formula character (=, +, -, @), prefix
    it with a single quote so Excel treats it as text, not a formula.
    """
    if val is None:
        return ""
    if isinstance(val, str) and val.startswith(_INJECTION_PREFIXES):
        return f"'{val}"
    return val


def export_xlsx(db: Database, output_path: str) -> str:
    rows = db.list_excel_view()
    wb = Workbook()
    ws = wb.active
    ws.title = "Game Library"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for r, row in enumerate(rows, start=2):
        for c, header in enumerate(HEADERS, 1):
            val = _sanitize_cell(row.get(header, ""))
            cell = ws.cell(row=r, column=c, value=val)
            if header == "SHORT DESCRIPTION":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

    for i, header in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(i)].width = _COL_WIDTHS[header]
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(out)
    except OSError as e:
        errno = getattr(e, "errno", None)
        if errno in (13, 5):
            raise PermissionError(
                f"Cannot write to '{out}'. The file may be open in another "
                f"program (e.g. Excel). Close it and try again."
            ) from e
        raise OSError(f"Could not write Excel file to '{out}': {e}") from e
    return str(out)
