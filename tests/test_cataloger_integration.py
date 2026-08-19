"""End-to-end pipeline tests with mocked API responses (no network).

These validate that RAWG/TheGamesDB JSON shapes are normalised correctly into
catalog columns and persisted to SQLite.
"""
from pathlib import Path

from playcache.cataloger import Cataloger
from playcache.config import Config
from playcache.db import Database
from playcache.models import GameRecord
from playcache.textutils import (
    best_match,
    clean_search_query,
    format_rating,
    join_names,
    strip_html,
    truncate,
)

# ---------- Canned RAWG responses (match the documented API shape) ---------- #
RAWG_DB = {
    11226: {
        "id": 11226, "slug": "hollow-knight", "name": "Hollow Knight",
        "released": "2017-02-24", "rating": 4.50, "ratings_count": 5000,
        "metacritic": 90,
        "description_raw": "A beautifully crafted action-adventure set in a vast "
                           "interconnected world. Explore twisting caverns and "
                           "fight corrupted creatures.",
        "genres": [{"id": 4, "name": "Action"}, {"id": 51, "name": "Indie"}],
        "developers": [{"id": 123, "name": "Team Cherry"}],
        "publishers": [{"id": 123, "name": "Team Cherry"}],
        "stores": [{"store": {"id": 1, "name": "Steam"}},
                   {"store": {"id": 6, "name": "GOG"}}],
        "background_image": "https://media.rawg.io/media/hollow.jpg",
        "website": "https://hollowknight.com",
    },
    31265: {
        "id": 31265, "slug": "deep-rock-galactic", "name": "Deep Rock Galactic",
        "released": "2020-05-13", "rating": 4.75, "ratings_count": 8000,
        "metacritic": 84,
        "description_raw": "Co-op shooter for 1-4 players; mine minerals and "
                           "fight alien bugs as dwarven space miners.",
        "genres": [{"id": 4, "name": "Action"}, {"id": 7, "name": "Shooter"}],
        "developers": [{"id": 200, "name": "Ghost Ship Games"}],
        "publishers": [{"id": 201, "name": "Coffee Stain Publishing"}],
        "stores": [{"store": {"id": 1, "name": "Steam"}}],
        "background_image": "https://media.rawg.io/media/drg.jpg",
        "website": "https://deeprockgalactic.com",
    },
}

_SEARCH_PROJECTION = (
    "id", "slug", "name", "released", "rating", "rating_top",
    "ratings_count", "metacritic", "background_image", "platforms", "stores",
)


class FakeRAWGClient:
    """A drop-in RAWGClient substitute returning canned data by query."""
    name = "rawg"
    request_count = 0

    def __init__(self, config):
        self.config = config

    def is_available(self):
        return True

    def search(self, query, page_size=5):
        q = query.lower()
        out = []
        for detail in RAWG_DB.values():
            name = detail["name"].lower()
            if q in name or any(w in name for w in q.split() if len(w) > 2):
                out.append({k: detail[k] for k in _SEARCH_PROJECTION if k in detail})
        return out

    def get_details(self, game_id):
        return dict(RAWG_DB[game_id])

    def fetch(self, record: GameRecord, *, overrides: dict | None = None) -> GameRecord:
        query = clean_search_query(record.game_name or record.folder_name)
        results = self.search(query)
        cand = best_match(query, ((r.get("name", ""), r) for r in results),
                          threshold=self.config.fuzzy_threshold)
        if not cand:
            record.fetch_status = "not_found"
            record.fetch_message = "no match"
            return record
        detail = self.get_details(cand.get("id"))
        record.game_name = detail.get("name") or record.game_name
        record.rawg_id = detail.get("id")
        record.rawg_slug = detail.get("slug")
        record.release_date = detail.get("released")
        record.developer = join_names(detail.get("developers"))
        record.publisher = join_names(detail.get("publishers"))
        record.game_type = join_names(detail.get("genres"))
        mc = detail.get("metacritic")
        record.metacritic_score = int(mc) if isinstance(mc, (int, float)) else None
        if detail.get("rating") and detail.get("ratings_count", 0) > 0:
            record.user_rating = format_rating(float(detail["rating"]) * 2.0)
        record.short_description = truncate(
            detail.get("description_raw") or strip_html(detail.get("description")),
            self.config.description_max_chars,
        )
        record.cover_url = detail.get("background_image")
        record.website = detail.get("website")
        record.fetch_status = "ok"
        record.data_source = self.name
        return record


class NotFoundRAWGClient(FakeRAWGClient):
    def search(self, query, page_size=5):
        return []


class FakeTGDBClient:
    name = "thegamesdb"
    request_count = 0

    def __init__(self, config):
        self.config = config

    def is_available(self):
        return True

    def fetch(self, record: GameRecord, *, overrides: dict | None = None) -> GameRecord:
        record.thegamesdb_id = 9999
        record.game_name = record.game_name or "Hollow Knight"
        record.short_description = "TGDB overview text."
        record.game_type = "Action / Platformer"
        record.developer = "Team Cherry"
        record.esrb_rating = "T - Teen"
        record.cover_url = "https://cdn.thegamesdb.net/images/large/boxart/front/9999-1.jpg"
        record.fetch_status = "ok"
        record.data_source = self.name
        return record


class NotFoundTGDBClient(FakeTGDBClient):
    def fetch(self, record: GameRecord, *, overrides: dict | None = None) -> GameRecord:
        record.fetch_status = "not_found"
        record.fetch_message = f"TGDB: no match for '{record.game_name}'"
        return record


def _make_tree(tmp_path: Path):
    (tmp_path / "Hollow Knight").mkdir()
    (tmp_path / "Deep Rock Galactic [SteamRip]").mkdir()


def test_pipeline_rawg_fetch_and_store(tmp_path):
    """RAWG primary with TGDB unavailable — user_rating/metacritic stored, esrb empty."""
    _make_tree(tmp_path)
    cfg = Config()
    cfg.db_path = str(tmp_path / "cat.db")
    db = Database(cfg.db_path)
    cat = Cataloger(cfg, db=db, rawg=FakeRAWGClient(cfg),
                    tgdb=NotFoundTGDBClient(cfg))

    summary = cat.scan_to_db(str(tmp_path))
    assert summary["ok"] == 2
    assert summary["stored"] == 2

    records = {r.folder_name: r for r in db.all_records()}
    for rec in records.values():
        assert rec.data_source == "rawg"
        assert rec.user_rating in ("9/10", "9.5/10")
        assert rec.metacritic_score in (84, 90)
        assert rec.esrb_rating == ""  # TGDB unavailable; RAWG has no ESRB
        assert rec.thegamesdb_id is None


def test_pipeline_rawg_merge_from_tgdb(tmp_path):
    """After RAWG succeeds, missing fields (esrb_rating) are filled from TGDB."""
    _make_tree(tmp_path)
    cfg = Config()
    cfg.db_path = str(tmp_path / "cat.db")
    db = Database(cfg.db_path)
    cat = Cataloger(cfg, db=db, rawg=FakeRAWGClient(cfg), tgdb=FakeTGDBClient(cfg))

    cat.scan_to_db(str(tmp_path))
    records = list(db.all_records())

    for rec in records:
        assert rec.data_source == "rawg"
        assert rec.user_rating in ("9/10", "9.5/10")
        assert rec.metacritic_score in (84, 90)
        assert rec.rawg_id in (11226, 31265)
        assert rec.esrb_rating == "T - Teen"  # merged from TGDB
        assert rec.thegamesdb_id == 9999  # merged from TGDB


def test_pipeline_fallback_to_tgdb(tmp_path):
    """When RAWG (primary) finds no match, fall back to TheGamesDB."""
    _make_tree(tmp_path)
    cfg = Config()
    cfg.db_path = str(tmp_path / "cat.db")
    db = Database(cfg.db_path)
    cat = Cataloger(cfg, db=db, rawg=NotFoundRAWGClient(cfg),
                    tgdb=FakeTGDBClient(cfg))

    summary = cat.scan_to_db(str(tmp_path))
    assert summary["ok"] == 2

    records = {r.folder_name: r for r in db.all_records()}
    hk = records["Hollow Knight"]
    assert hk.data_source == "thegamesdb"
    assert hk.esrb_rating == "T - Teen"
    assert hk.thegamesdb_id == 9999
    assert hk.game_type == "Action / Platformer"
    assert hk.short_description == "TGDB overview text."
    assert hk.user_rating == ""  # TGDB has no numeric rating


def test_pipeline_skips_already_catalogued(tmp_path):
    _make_tree(tmp_path)
    cfg = Config()
    cfg.db_path = str(tmp_path / "cat.db")
    db = Database(cfg.db_path)
    cat = Cataloger(cfg, db=db, rawg=FakeRAWGClient(cfg), tgdb=FakeTGDBClient(cfg))

    cat.scan_to_db(str(tmp_path))
    summary = cat.scan_to_db(str(tmp_path), rescan=False)
    assert summary["skipped"] == 2
    assert summary["ok"] == 0
    assert summary["stored"] == 0


def test_excel_export_round_trip(tmp_path):
    import openpyxl

    from playcache.exporter import export_xlsx

    _make_tree(tmp_path)
    cfg = Config()
    cfg.db_path = str(tmp_path / "cat.db")
    db = Database(cfg.db_path)
    cat = Cataloger(cfg, db=db, rawg=FakeRAWGClient(cfg), tgdb=FakeTGDBClient(cfg))
    cat.scan_to_db(str(tmp_path))

    out = export_xlsx(db, str(tmp_path / "out.xlsx"))
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == ["GAME NAME", "PLATFORM", "GOG / STEAM",
                       "USER RATING", "GAME TYPE", "SHORT DESCRIPTION"]
    assert ws.max_row == 1 + db.count()
