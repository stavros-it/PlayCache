"""Tests for the Excel exporter (formula injection, layout, error handling)."""
from pathlib import Path

import pytest
from openpyxl import load_workbook

from playcache.db import Database
from playcache.exporter import export_xlsx
from playcache.models import GameRecord


def _sample(folder_path="/games/Test", **kw):
    base = {
        "folder_name": "Test",
        "folder_path": folder_path,
        "game_name": "Test Game",
        "platform": "PC",
        "store": "GOG",
        "user_rating": "9/10",
        "game_type": "Action",
        "short_description": "A test game.",
        "rawg_id": 1234,
        "rawg_slug": "test-game",
        "release_date": "2020-01-01",
        "developer": "TestDev",
        "publisher": "TestPub",
        "metacritic_score": 85,
        "data_source": "rawg",
        "fetch_status": "ok",
        "esrb_rating": "T - Teen",
        "manual_overrides": "",
    }
    base.update(kw)
    return GameRecord(**base)


def test_export_creates_xlsx(tmp_path):
    db = Database(str(tmp_path / "test.db"))
    db.upsert(_sample())
    out = export_xlsx(db, str(tmp_path / "out.xlsx"))
    assert Path(out).is_file()


def test_export_formula_injection_sanitized(tmp_path):
    """A game name starting with = must be prefixed to prevent formula execution."""
    db = Database(str(tmp_path / "test.db"))
    db.upsert(_sample(game_name='=cmd|"/c calc"!A1'))
    out = export_xlsx(db, str(tmp_path / "out.xlsx"))
    wb = load_workbook(out)
    ws = wb.active
    val = ws.cell(row=2, column=1).value
    assert val.startswith("'=")


def test_export_plus_injection_sanitized(tmp_path):
    db = Database(str(tmp_path / "test.db"))
    db.upsert(_sample(game_name="+cmd"))
    out = export_xlsx(db, str(tmp_path / "out.xlsx"))
    wb = load_workbook(out)
    ws = wb.active
    val = ws.cell(row=2, column=1).value
    assert val.startswith("'+")


def test_export_at_injection_sanitized(tmp_path):
    db = Database(str(tmp_path / "test.db"))
    db.upsert(_sample(short_description="@SUM(1+1)"))
    out = export_xlsx(db, str(tmp_path / "out.xlsx"))
    wb = load_workbook(out)
    ws = wb.active
    val = ws.cell(row=2, column=6).value
    assert val.startswith("'@")


def test_export_normal_names_not_modified(tmp_path):
    db = Database(str(tmp_path / "test.db"))
    db.upsert(_sample(game_name="Half-Life"))
    out = export_xlsx(db, str(tmp_path / "out.xlsx"))
    wb = load_workbook(out)
    ws = wb.active
    val = ws.cell(row=2, column=1).value
    assert val == "Half-Life"


def test_export_creates_parent_dir(tmp_path):
    db = Database(str(tmp_path / "test.db"))
    db.upsert(_sample())
    nested = tmp_path / "nested" / "deep" / "out.xlsx"
    out = export_xlsx(db, str(nested))
    assert Path(out).is_file()


def test_export_permission_error_message(tmp_path):
    db = Database(str(tmp_path / "test.db"))
    db.upsert(_sample())
    with pytest.raises(PermissionError, match="may be open in another"):
        export_xlsx(db, str(tmp_path))


def test_export_empty_db(tmp_path):
    db = Database(str(tmp_path / "test.db"))
    out = export_xlsx(db, str(tmp_path / "empty.xlsx"))
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "GAME NAME"
    assert ws.cell(row=2, column=1).value is None
